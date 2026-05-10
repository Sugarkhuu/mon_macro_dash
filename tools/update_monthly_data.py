"""Update monthly source data from public data providers.

Current coverage:
    - CPI headline index from NSO PxWeb table DT_NSO_0600_001V3
    - CPI month-over-month group changes from NSO PxWeb table DT_NSO_0600_009V1

Examples:
    python tools/update_monthly_data.py --source cpi --latest --dry-run
    python tools/update_monthly_data.py --source cpi --month 2026-04
"""

from __future__ import annotations

import argparse
import json
import sys
from dataclasses import dataclass
from pathlib import Path
from typing import Any

import requests
from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
MONTHLY_WORKBOOK = ROOT / "data" / "data_monthly.xlsx"
PXWEB_BASE = "https://data.1212.mn/api/v1/en/NSO/Economy%2C%20environment/Consumer%20Price%20Index"
HEADLINE_CPI_TABLE = f"{PXWEB_BASE}/DT_NSO_0600_001V3.px"
CPI_MOM_TABLE = f"{PXWEB_BASE}/DT_NSO_0600_009V1.px"
REFERENCE_YEAR_TEXT = "2023=100"


CPI_MOM_COLUMNS = {
    "0": "cpi_20_mom",
    "1": "cpi_food_20_mom",
    "2": "cpi_alco_20_mom",
    "3": "cpi_clot_20_mom",
    "4": "cpi_elec_20_mom",
    "5": "cpi_furn_20_mom",
    "6": "cpi_heal_20_mom",
    "7": "cpi_tran_20_mom",
    "8": "cpi_comm_20_mom",
    "9": "cpi_recr_20_mom",
    "10": "cpi_educ_20_mom",
    "11": "cpi_hote_20_mom",
    "12": "cpi_insu_20_mom",
    "13": "cpi_oth_20_mom",
}


@dataclass
class PxVariable:
    code: str
    text: str
    values: list[str]
    value_texts: list[str]


def main() -> int:
    parser = argparse.ArgumentParser(description="Append public monthly data into data_monthly.xlsx.")
    parser.add_argument("--source", choices=["cpi"], default="cpi")
    parser.add_argument(
        "--month",
        action="append",
        help="Month to update in YYYY-MM format. Can be repeated.",
    )
    parser.add_argument(
        "--latest",
        action="store_true",
        help="Update the latest month available in the source.",
    )
    parser.add_argument(
        "--workbook",
        default=str(MONTHLY_WORKBOOK),
        help="Path to data_monthly.xlsx.",
    )
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="Show what would change without saving the workbook.",
    )
    args = parser.parse_args()

    if args.source == "cpi":
        changes = update_cpi(
            workbook_path=Path(args.workbook),
            requested_months=args.month,
            latest=args.latest,
            dry_run=args.dry_run,
        )
        print_change_summary(changes, args.dry_run)
        return 0

    raise ValueError(f"Unsupported source: {args.source}")


def update_cpi(
    workbook_path: Path,
    requested_months: list[str] | None,
    latest: bool,
    dry_run: bool,
) -> list[str]:
    if not workbook_path.exists():
        raise FileNotFoundError(workbook_path)

    headline_meta = get_px_metadata(HEADLINE_CPI_TABLE)
    mom_meta = get_px_metadata(CPI_MOM_TABLE)

    months = resolve_months(headline_meta, requested_months, latest)
    reference_value = value_for_text(headline_meta, "Reference year", REFERENCE_YEAR_TEXT)
    headline_group = value_for_text(headline_meta, "Group", "Overall  index")

    month_values_headline = values_for_months(headline_meta, months)
    month_values_mom = values_for_months(mom_meta, months)
    mom_reference_value = value_for_text(mom_meta, "Reference year", REFERENCE_YEAR_TEXT)

    headline_rows = fetch_px_json(
        HEADLINE_CPI_TABLE,
        headline_meta,
        {
            "Reference year": [reference_value],
            "Group": [headline_group],
            "Month": month_values_headline,
        },
    )
    mom_rows = fetch_px_json(
        CPI_MOM_TABLE,
        mom_meta,
        {
            "Reference year": [mom_reference_value],
            "Group": list(CPI_MOM_COLUMNS),
            "Month": month_values_mom,
        },
    )

    updates = build_cpi_updates(headline_rows, mom_rows, headline_meta, mom_meta)

    workbook = load_workbook(workbook_path)
    if "data" not in workbook.sheetnames:
        raise ValueError(f"{workbook_path} does not contain a 'data' sheet.")
    sheet = workbook["data"]

    header = header_map(sheet)
    row_lookup = row_map(sheet)

    changes: list[str] = []
    for period, period_updates in sorted(updates.items()):
        row_index = row_lookup.get(period)
        if row_index is None:
            row_index = sheet.max_row + 1
            sheet.cell(row=row_index, column=1, value=period)
            row_lookup[period] = row_index
            changes.append(f"add row {period}")

        for column_name, value in period_updates.items():
            column_index = header.get(column_name)
            if column_index is None:
                changes.append(f"skip {period}.{column_name}: column missing")
                continue

            cell = sheet.cell(row=row_index, column=column_index)
            if cell.value not in (None, ""):
                changes.append(f"keep {period}.{column_name}: existing={cell.value}")
                continue

            cell.value = value
            changes.append(f"set {period}.{column_name}={value}")

    if not dry_run:
        workbook.save(workbook_path)

    return changes


def get_px_metadata(url: str) -> dict[str, PxVariable]:
    response = requests.get(url, timeout=30)
    response.raise_for_status()
    payload = response.json()
    variables = {}
    for variable in payload["variables"]:
        text = str(variable["text"])
        variables[text] = PxVariable(
            code=str(variable["code"]),
            text=text,
            values=[str(value) for value in variable["values"]],
            value_texts=[str(value) for value in variable.get("valueTexts", [])],
        )
    return variables


def resolve_months(
    metadata: dict[str, PxVariable],
    requested_months: list[str] | None,
    latest: bool,
) -> list[str]:
    source_months = metadata["Month"].value_texts
    if requested_months:
        months = requested_months
    elif latest:
        months = [source_months[0]]
    else:
        raise ValueError("Specify --month YYYY-MM or --latest.")

    missing = [month for month in months if month not in source_months]
    if missing:
        raise ValueError(f"Month(s) not found in source: {', '.join(missing)}")
    return months


def value_for_text(metadata: dict[str, PxVariable], variable_text: str, value_text: str) -> str:
    variable = metadata[variable_text]
    normalized_target = normalize_label(value_text)
    for value, text in zip(variable.values, variable.value_texts):
        if normalize_label(text) == normalized_target:
            return value
    raise ValueError(f"Could not find {value_text!r} in {variable_text}.")


def values_for_months(metadata: dict[str, PxVariable], months: list[str]) -> list[str]:
    variable = metadata["Month"]
    lookup = {text: value for value, text in zip(variable.values, variable.value_texts)}
    return [lookup[month] for month in months]


def fetch_px_json(
    url: str,
    metadata: dict[str, PxVariable],
    selections: dict[str, list[str]],
) -> list[dict[str, Any]]:
    query = {"query": [], "response": {"format": "JSON"}}
    for variable_text, variable in metadata.items():
        selected_values = selections[variable_text]
        query["query"].append(
            {
                "code": variable.code,
                "selection": {
                    "filter": "item",
                    "values": selected_values,
                },
            }
        )

    response = requests.post(
        url,
        data=json.dumps(query).encode("utf-8"),
        headers={"Content-Type": "application/json"},
        timeout=30,
    )
    response.raise_for_status()
    return response.json()["data"]


def build_cpi_updates(
    headline_rows: list[dict[str, Any]],
    mom_rows: list[dict[str, Any]],
    headline_meta: dict[str, PxVariable],
    mom_meta: dict[str, PxVariable],
) -> dict[str, dict[str, float]]:
    headline_months = value_to_text(headline_meta["Month"])
    mom_months = value_to_text(mom_meta["Month"])
    updates: dict[str, dict[str, float]] = {}

    for row in headline_rows:
        _reference, _group, month_value = row["key"]
        period = month_to_period(headline_months[month_value])
        value = parse_px_number(row["values"][0])
        if value is not None:
            updates.setdefault(period, {})["cpi"] = value

    for row in mom_rows:
        _reference, group_value, month_value = row["key"]
        column_name = CPI_MOM_COLUMNS.get(group_value)
        if not column_name:
            continue
        period = month_to_period(mom_months[month_value])
        value = parse_px_number(row["values"][0])
        if value is not None:
            updates.setdefault(period, {})[column_name] = value

    return updates


def value_to_text(variable: PxVariable) -> dict[str, str]:
    return dict(zip(variable.values, variable.value_texts))


def month_to_period(month: str) -> str:
    year, month_number = month.split("-", 1)
    return f"{year[-2:]}M{int(month_number):02d}"


def parse_px_number(value: Any) -> float | None:
    if value in (None, "", "..", "..."):
        return None
    return float(str(value).replace(",", ""))


def header_map(sheet: Any) -> dict[str, int]:
    return {
        str(cell.value): cell.column
        for cell in sheet[1]
        if cell.value not in (None, "")
    }


def row_map(sheet: Any) -> dict[str, int]:
    rows = {}
    for row_index in range(2, sheet.max_row + 1):
        value = sheet.cell(row=row_index, column=1).value
        if value not in (None, ""):
            rows[str(value)] = row_index
    return rows


def normalize_label(value: str) -> str:
    return " ".join(value.strip().lower().split())


def print_change_summary(changes: list[str], dry_run: bool) -> None:
    label = "DRY RUN" if dry_run else "UPDATED"
    print(f"{label}: {len(changes)} action(s)")
    for change in changes:
        print(f"- {change}")


if __name__ == "__main__":
    try:
        raise SystemExit(main())
    except Exception as exc:
        print(f"ERROR: {exc}", file=sys.stderr)
        raise SystemExit(1)
